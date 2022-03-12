from numpy import load, linspace
from pickle import load as pload
# noinspection PyPackageRequirements
from openpyxl import Workbook

from dsw import Monitor


if __name__ == "__main__":
    book, monitor = Workbook(), Monitor()

    data = load(file="./results/data/step_3_stability_evaluation.npy")

    sheet = book.create_sheet(title="stability task", index=0)
    sheet.append(["coding algorithm", "constraint set",
                  "size of digital data (bit)", "used nucleotides (nt)", "code rate"])
    algorithm_labels = ["DNA Fountain", "Yin-Yang Code", "HEDGES", "SPIDER-WEB"]

    print("Load the stability test.")
    current = 0
    for sample in data:
        row_data = [algorithm_labels[sample[0]], sample[1],
                    8 * 1024 * 64, sample[2]]
        if sample[2] > 0:
            row_data.append((8 * 1024 * 64) / sample[2])
        else:
            row_data.append("N/A")
        sheet.append(row_data)

        monitor.output(current + 1, 388800)
        current += 1

    with open("./results/data/step_4_repairability_multiple_errors.pkl", "rb") as file:
        task_1, task_2, task_3 = pload(file)

    sheet = book.create_sheet(title="repairability task 1", index=1)
    sheet.append(["constraint set", "introduced error number", "aligned error rate", "variation coefficient",
                  "detection flag (search only)", "detection flag (VT code)", "detection flag (combined)",
                  "number of solutions (search only)", "number of solutions (combined)", "repair flag"])

    print("Load the repairability task 1.")
    current = 0
    for (filter_index, error_time), values in task_1.items():
        for sample in values:
            sheet.append([filter_index, error_time, sample[0], sample[1],
                          bool(sample[2]), bool(sample[4]), bool(sample[5]),
                          int(sample[3]), int(sample[6]), bool(sample[7])])
            monitor.output(current + 1, 96000)
            current += 1

    sheet = book.create_sheet(title="repairability task 2", index=2)
    sheet.append(["constraint set", "introduced error number", "aligned error rate", "variation coefficient",
                  "detection flag (search only)", "detection flag (VT code)", "detection flag (combined)",
                  "number of solutions (search only)", "number of solutions (combined)", "repair flag"])

    print("Load the repairability task 2.")
    current = 0
    for (filter_index, error_time), values in task_2.items():
        for sample in values:
            sheet.append([filter_index, error_time, sample[0], sample[1],
                          bool(sample[2]), bool(sample[4]), bool(sample[5]),
                          int(sample[3]), int(sample[6]), bool(sample[7])])
            monitor.output(current + 1, 32000)
            current += 1

    sheet = book.create_sheet(title="repairability task 3", index=3)
    sheet.append(["DNA string length", "introduced error time",
                  "average correction rate", "redundancy proportion", "average runtime (second)"])

    print("Load the repairability task 3.")
    current, observed_length = 0, 10
    for length_index, length in enumerate(linspace(100, 400, 26)):
        for error_index, error_time in enumerate([1, 2, 3, 4]):
            rate = task_3[0, error_index, length_index]
            runtime = task_3[1, error_index, length_index]
            redundancy = (observed_length + 1) + (length + observed_length + 1) * (1 - rate) / rate
            sheet.append([length, error_time,
                          rate, redundancy / (length + length), runtime])
            monitor.output(current + 1, 104)
            current += 1

    reconstructions = load(file="./results/data/step_5_encrypability_reconstruction.npy")
    sheet = book.create_sheet(title="encryption task", index=4)
    sheet.append(["constraint set", "repeats"] + linspace(0, 10000, 10001, dtype=int).tolist())

    print("Load the encryption task.")
    current = 0
    for filter_index, data in enumerate(reconstructions):
        for repeat, iteration_data in enumerate(data):
            sheet.append([filter_index + 1, repeat + 1] + iteration_data.tolist())
            monitor.output(current + 1, 1200)
            current += 1

    book.save("./results/raw data.xlsx")
