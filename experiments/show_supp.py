from collections import Counter
from logging import getLogger, CRITICAL
from matplotlib import pyplot, rcParams, patches
from numpy import array, arange, zeros, ones, zeros_like, linspace, random, median, where
from numpy import min, max, mean, sum, std, abs, sqrt, log2, log10, percentile, polyfit
from openpyxl import Workbook
from scipy.stats import gaussian_kde
from warnings import filterwarnings

from dsw import obtain_vertices

from experiments import load_data, local_bio_filters

filterwarnings("ignore")

getLogger("matplotlib").setLevel(CRITICAL)

rcParams["font.family"] = "Arial"
rcParams["mathtext.fontset"] = "custom"
rcParams["mathtext.rm"] = "Linux Libertine"
rcParams["mathtext.cal"] = "Lucida Calligraphy"
rcParams["mathtext.it"] = "Linux Libertine:italic"
rcParams["mathtext.bf"] = "Linux Libertine:bold"


def supp01():
    capacities, _ = load_data(load_path="./raw/capacity_reference.pkl")

    book = Workbook()
    sheet = book.create_sheet(title="supp01", index=0)
    sheet.append(["set index", "homopolymer run-length", "regionalized GC content", "undesired motifs",
                  "capacity"])

    for constraint_index, bio_filter in local_bio_filters.items():
        save_data = [constraint_index]
        row_info = constraint_index.ljust(20)
        if bio_filter.max_homopolymer_runs is not None:
            row_info += " | " + str(bio_filter.max_homopolymer_runs).ljust(20)
            save_data.append(str(bio_filter.max_homopolymer_runs))
        else:
            row_info += " | " + "N/A".ljust(20)
            save_data.append("N/A")

        if bio_filter.gc_range is not None:
            if bio_filter.gc_range[0] != bio_filter.gc_range[1]:
                gc_range = [str(int(bio_filter.gc_range[0] * 100)) + "%", str(int(bio_filter.gc_range[1] * 100)) + "%"]
                save_data.append(gc_range[0] + " ~ " + gc_range[1])
                row_info += " | " + (gc_range[0] + " ~ " + gc_range[1]).ljust(25)
            else:
                row_info += " | " + (str(int(bio_filter.gc_range[0] * 100)) + "%").ljust(25)
                save_data.append(str(int(bio_filter.gc_range[0] * 100)) + "%")
        else:
            row_info += " | " + "N/A".ljust(25)
            save_data.append("N/A")

        if bio_filter.undesired_motifs is not None:
            row_info += " | " + str(bio_filter.undesired_motifs)[1:-1].replace("\"", "").replace("\'", "").ljust(100)
            save_data.append(str(bio_filter.undesired_motifs)[1:-1].replace("\"", "").replace("\'", ""))
        else:
            row_info += " | " + "N/A".ljust(100)
            save_data.append("N/A")

        row_info += " | " + "%.4f" % capacities[int(constraint_index) - 1]
        save_data.append("%.4f" % capacities[int(constraint_index) - 1])

        sheet.append(save_data)

    book.save("./show/supp01.xlsx")


def supp02():
    filter_indices = ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12"]
    colors = pyplot.get_cmap("Set3")(linspace(0, 1, 12))
    record = load_data(load_path="./raw/generation_evaluation.pkl")

    pyplot.figure(figsize=(10, 4), tight_layout=True)

    ax = pyplot.subplot(1, 1, 1)
    time_data = record["time"]

    for index in range(12):
        value = time_data[index]
        pyplot.bar([index], [value], fc=colors[index], ec="k", lw=0.75)
        pyplot.text(index, value + 2, "%.2f" % value, va="center", ha="center", fontsize=10)
    pyplot.xlabel("set index", fontsize=10)
    pyplot.ylabel("generation runtime (seconds)", fontsize=10)
    pyplot.xticks(range(12), filter_indices, fontsize=10)
    pyplot.yticks(arange(0, 121, 20), arange(0, 121, 20), fontsize=10)
    pyplot.xlim(-0.6, 11.6)
    pyplot.ylim(0, 120)
    # noinspection PyUnresolvedReferences
    ax.spines["top"].set_visible(False)
    # noinspection PyUnresolvedReferences
    ax.spines["right"].set_visible(False)

    pyplot.savefig("./show/supp02.pdf", format="pdf", bbox_inches="tight", dpi=600)
    pyplot.close()


def supp03():
    filter_indices = ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12"]
    colors = pyplot.get_cmap("Set3")(linspace(0, 1, 12))
    record = load_data(load_path="./raw/generation_evaluation.pkl")

    pyplot.figure(figsize=(10, 4), tight_layout=True)

    change_data = record["change"]

    ax = pyplot.subplot(1, 1, 1)

    for index in range(12):
        value = change_data[str(index + 1).zfill(2)][0]
        pyplot.bar([index], [log2(value) / log2(4)], fc=colors[index], ec="k", lw=0.75)
        pyplot.text(index, log2(value) / log2(4) + 0.15, str(value), va="center", ha="center", fontsize=10)
    pyplot.xlabel("set index", fontsize=10)
    pyplot.ylabel("remain vertices after screening", fontsize=10)
    pyplot.xticks(range(12), filter_indices, fontsize=10)
    pyplot.yticks(arange(6, 11), ["4^" + str(v).zfill(2) for v in arange(6, 11)], fontsize=10)
    pyplot.xlim(-0.6, 11.6)
    pyplot.ylim(6, 10)
    # noinspection PyUnresolvedReferences
    ax.spines["top"].set_visible(False)
    # noinspection PyUnresolvedReferences
    ax.spines["right"].set_visible(False)

    pyplot.savefig("./show/supp03.pdf", format="pdf", bbox_inches="tight", dpi=600)
    pyplot.close()


def supp04():
    record = load_data(load_path="./raw/generation_evaluation.pkl")
    change_data = record["change"]

    book = Workbook()
    sheet = book.create_sheet(title="supp04", index=0)
    sheet.append(["trimming cycle",
                  "C01", "C02", "C03", "C04", "C05", "C06", "C07", "C08", "C09", "C10", "C11", "C12"])
    for cycle_index in range(13):
        values = [cycle_index]
        for set_index in range(12):
            numbers = change_data[str(set_index + 1).zfill(2)][1]
            if len(numbers) > cycle_index:
                values.append(numbers[cycle_index])
            else:
                values.append("")
        sheet.append(values)

    book.save("./show/supp04.xlsx")


def supp05():
    random_seed, param_number = 2021, 100

    random.seed(random_seed)
    fountain_params = random.uniform(low=array([0.1, 0.01]), high=array([1, 0.1]), size=(param_number, 2)).tolist()

    random.seed(random_seed)
    total_indices = array(list(range(6144)))
    random.shuffle(total_indices)
    yinyang_params = total_indices[:param_number]
    yinyang_params = yinyang_params.tolist()

    hedges_params = []
    correct_penalties = [-0.035, -0.082, -0.127, -0.229, -0.265, -0.324]
    for pattern_index, correct_penalty in enumerate(correct_penalties):
        hedges_params.append([pattern_index + 1, correct_penalty])
    hedges_params = array(hedges_params).tolist()

    book = Workbook()
    sheet = book.create_sheet(title="supp05", index=0)
    sheet.append(["param index", "DNA Fountain", "", "Yin-Yang Code", "HEDGES", ""])
    sheet.append(["", "c", "delta", "rule id", "pattern", "correct penalty"])
    for param_index in range(param_number):
        record = [str(param_index + 1)]
        for value in fountain_params[param_index]:
            record.append("%.2f" % value)
        record.append(str(yinyang_params[param_index]))
        if param_index < 6:
            record.append(str(int(hedges_params[param_index][0])))
            record.append("%.3f" % hedges_params[param_index][1])
        else:
            record.append("N/A")
            record.append("N/A")

        sheet.append(record)

    book.save("./show/supp05.xlsx")


def supp06():
    random_seed, param_number = 2021, 100

    spiderweb_params = []
    coding_graphs = load_data(load_path="./raw/graph_coding.pkl")
    for filter_index in ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12"]:
        vertices = obtain_vertices(accessor=coding_graphs[filter_index])
        random.seed(random_seed)
        random.shuffle(vertices)
        chosen_indices = vertices[:param_number]
        spiderweb_params.append(chosen_indices)
    spiderweb_params = array(spiderweb_params).T.tolist()

    book = Workbook()
    sheet = book.create_sheet(title="supp08", index=0)
    sheet.append(["param index", "SPIDER-WEB", "", "", "", "", "", "", "", "", "", "", ""])
    sheet.append(["", "C01", "C02", "C03", "C04", "C05", "C06", "C07", "C08", "C09", "C10", "C11", "C12"])
    for param_index in range(param_number):
        record = [str(param_index + 1)]
        for vertex_index in spiderweb_params[param_index]:
            record.append(str(vertex_index))

        sheet.append(record)

    book.save("./show/supp06.xlsx")


def supp07():
    display_data = [[[] for _ in range(4)] for _ in range(12)]
    record = load_data(load_path="./raw/coding_evaluation_1.pkl")
    for algorithm_index, filter_index, _, _, code_rate in record:
        display_data[int(filter_index - 1)][int(algorithm_index - 1)].append(code_rate)
    book = Workbook()
    sheet = book.create_sheet(title="supp07", index=0)
    sheet.append(["set index", "DNA Fountain", "Yin-Yang Code", "HEDGES", "SPIDER-WEB"])
    for index, data in enumerate(display_data):
        record = [str(index + 1).zfill(2)]
        for method_index, sub_data in enumerate(data):
            if sum(sub_data) == -len(sub_data):
                record.append("N/A")
            elif max(sub_data) - min(sub_data) < 0.001:
                record.append("%.2f" % mean(sub_data))
            else:
                sub_data = array(sub_data)
                sub_data = sub_data[sub_data > 0]
                record.append("%.2f ~ %.2f" % (min(sub_data), max(sub_data)))
        sheet.append(record)

    book.save("./show/supp07.xlsx")


def supp08():
    display_data = [[[] for _ in range(4)] for _ in range(12)]
    record = load_data(load_path="./raw/coding_evaluation_1.pkl")
    for algorithm_index, filter_index, _, _, code_rate in record:
        display_data[int(filter_index) - 1][int(algorithm_index) - 1].append(code_rate)

    book = Workbook()
    sheet = book.create_sheet(title="supp08", index=0)
    sheet.append(["set index", "DNA Fountain", "Yin-Yang Code", "HEDGES", "SPIDER-WEB"])
    for index, data in enumerate(display_data):
        record = [str(index + 1).zfill(2)]
        for method_index, sub_data in enumerate(data):
            if sum(sub_data) == -len(sub_data):
                record.append("N/A")
            elif max(sub_data) - min(sub_data) < 0.001:
                record.append("%.2f" % std(sub_data))
            else:
                sub_data = array(sub_data)
                sub_data = sub_data[sub_data > 0]
                record.append("%.4f" % std(sub_data))
        sheet.append(record)

    book.save("./show/supp08.xlsx")


def supp09():
    counts = ones(shape=(12, 6)) * 100
    record = load_data(load_path="./raw/coding_evaluation_1.pkl")
    for algorithm_index, filter_index, param_index, _, code_rate in record:
        if algorithm_index == 3 and code_rate == -1:  # HEDGES
            counts[int(filter_index) - 1, int(param_index) - 1] -= 1
    counts /= 100.0

    book = Workbook()
    sheet = book.create_sheet(title="supp09", index=0)
    sheet.append(["set index", "pattern 1", "pattern 2", "pattern 3", "pattern 4", "pattern 5", "pattern 6"])
    for index, data in enumerate(counts):
        record = [str(index + 1).zfill(2)]
        for value in data:
            record.append(str(int(value * 100)) + "%")
        sheet.append(record)

    book.save("./show/supp09.xlsx")


def supp10():
    filter_indices = ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12"]

    capacities, _ = load_data(load_path="./raw/capacity_reference.pkl")
    code_results = load_data(load_path="./raw/coding_evaluation_2.pkl")[0]

    pyplot.figure(figsize=(10, 8))
    for index, filter_index in enumerate(filter_indices):
        capacity, code_rates = capacities[index], code_results[index]
        pyplot.text(x=index, y=capacity + 0.01, s="%.4f" % capacity, ha="center", va="bottom", fontsize=9)
        pyplot.hlines(capacity, index - 0.4, index + 0.4, color="#81B8DF", linewidth=1)
        if max(code_rates) - min(code_rates) < 0.005 or max(code_rates) > capacity:
            code_rate = median(code_rates)
            if code_rate > capacity:
                code_rate = capacity
            pyplot.hlines(code_rate, index - 0.25, index + 0.25, linewidths=1, edgecolors="#FE817D", zorder=3)
            pyplot.scatter([index], code_rate, color="white", edgecolor="#FE817D", linewidth=1, s=15, zorder=4)
        else:
            violin = pyplot.violinplot(dataset=code_rates, positions=[index], bw_method=0.5, showextrema=False)
            for patch in violin["bodies"]:
                patch.set_edgecolor("#FE817D")
                patch.set_facecolor("#FCBBAE")
                patch.set_linewidth(1)
                patch.set_alpha(1)
            pyplot.scatter([index], median(code_rates), color="white", edgecolor="#FE817D",
                           linewidth=1, s=15, zorder=4)
        if index % 2 != 0:
            pyplot.fill_between([index - 0.5, index + 0.5], [0.92, 0.92], [2.08, 2.08], color="#F1F1F1", zorder=0)

    pyplot.xlabel("set index", fontsize=10)
    pyplot.xticks(range(12), filter_indices, fontsize=10)
    pyplot.xlim(-0.5, 11.5)
    pyplot.ylabel("code rate", fontsize=10)
    pyplot.yticks([1.0, 1.1, 1.2, 1.3, 1.4, 1.5, 1.6, 1.7, 1.8, 1.9, 2.0],
                  ["1.0", "1.1", "1.2", "1.3", "1.4", "1.5", "1.6", "1.7", "1.8", "1.9", "2.0"], fontsize=10)
    pyplot.ylim(0.95, 2.05)
    pyplot.savefig("./show/supp10.pdf", format="pdf", bbox_inches="tight", dpi=600)
    pyplot.close()


def supp11():
    data = load_data(load_path="./raw/correction_evaluation_1.pkl")

    pyplot.figure(figsize=(10, 6), tight_layout=True)

    ax = pyplot.subplot(1, 1, 1)
    detection_rates, correction_rates = [], []
    for index, values in enumerate(data.values()):
        counts = values[0][:, 3]
        counts = counts[where(values[0][:, 4] > 0)]
        detection_rates.append(len(where(values[0][:, 0] == index)[0]) / 10000.0)
        correction_rates.append(len(counts[counts == 1]) / 10000.0)

    pyplot.bar(arange(len(detection_rates))[:14], detection_rates[:14],
               fc="w", ec="k", lw=0.75, label="for detection")
    pyplot.bar(arange(len(correction_rates))[:14], correction_rates[:14],
               fc="#A9D18E", ec="k", lw=0.75, label="for correction")
    for index, (detection_rate, correction_rate) in enumerate(zip(detection_rates[:15], correction_rates[:15])):
        if detection_rate > 0.05:
            pyplot.text(index, detection_rate + 0.01, ("%.1f" % (detection_rate * 100)) + "%",
                        va="bottom", ha="center", fontsize=9)
        else:
            pyplot.scatter([index], [0.02], marker="x", s=30, color="k", zorder=3)

        if correction_rate > 0.05:
            pyplot.text(index, correction_rate - 0.01, ("%.1f" % (correction_rate * 100)) + "%",
                        va="top", ha="center", fontsize=9)
        else:
            pyplot.scatter([index], [0.02], marker="x", s=30, color="k", zorder=3)

    pyplot.legend(loc="upper right", fontsize=9, framealpha=1)
    pyplot.xlabel("error rate", fontsize=10)
    pyplot.ylabel("success rate", fontsize=10)
    pyplot.xticks(arange(0, 15), ["%.1f" % (value / 200.0 * 100.0) + "%" for value in arange(0, 15)], fontsize=10)
    pyplot.yticks(linspace(0, 1, 6), [str(int(v * 100)) + "%" for v in linspace(0, 1, 6)], fontsize=10)
    pyplot.xlim(-0.6, 14.6)
    pyplot.ylim(0.00, 1.05)
    # noinspection PyUnresolvedReferences
    ax.spines["top"].set_visible(False)
    # noinspection PyUnresolvedReferences
    ax.spines["right"].set_visible(False)

    pyplot.savefig("./show/supp11.pdf", format="pdf", bbox_inches="tight", dpi=600)
    pyplot.close()


def supp12():
    data = load_data(load_path="./raw/correction_evaluation_1.pkl")

    pyplot.figure(figsize=(10, 6), tight_layout=True)
    for location, error_rate in enumerate([0.005, 0.010, 0.015, 0.020, 0.025, 0.030, 0.035, 0.040]):
        values = data[error_rate]
        original, combined = [], []
        for value in values[0]:
            if value[0] > 0 and value[3] > 0:
                original.append(value[2])
                combined.append(value[3])
        original, combined = array(original), array(combined)
        mean_value, max_value = mean(original), max(original)
        pyplot.bar([location - 0.2], [log2(mean_value) + 0.5], bottom=-0.5, width=0.4, fc="#FE817D", ec="k", lw=0.75)
        pyplot.vlines(location - 0.2, log2(mean_value), log2(max_value), lw=0.75, color="k")
        pyplot.hlines(log2(max_value), location - 0.3, location - 0.1, lw=0.75, color="k")
        pyplot.text(location - 0.2, log2(max_value) + 0.15, "%.4f" % max_value,
                    va="center", ha="center", fontsize=8.5)
        pyplot.text(location - 0.2, log2(mean_value) - 0.15, "%.4f" % mean_value,
                    va="center", ha="center", fontsize=8.5)
        mean_value, max_value = mean(combined), max(combined)
        pyplot.bar([location + 0.2], [log2(mean_value) + 0.5], bottom=-0.5, width=0.4, fc="#00C382", ec="k", lw=0.75)
        pyplot.vlines(location + 0.2, log2(mean_value), log2(max_value), lw=0.75, color="k")
        pyplot.hlines(log2(max_value), location + 0.1, location + 0.3, lw=0.75, color="k")
        pyplot.text(location + 0.2, log2(max_value) + 0.15, "%.4f" % max_value,
                    va="center", ha="center", fontsize=8.5)
        pyplot.text(location + 0.2, log2(mean_value) - 0.15, "%.4f" % mean_value,
                    va="center", ha="center", fontsize=8.5)

    handles = [patches.Patch(facecolor="#FE817D", edgecolor="black", linewidth=0.75, label="search-only"),
               patches.Patch(facecolor="#00C382", edgecolor="black", linewidth=0.75, label="combined")]
    pyplot.legend(loc="upper left", handles=handles, fontsize=9)
    pyplot.xlabel("error rate", fontsize=10)
    pyplot.ylabel("candidate number", fontsize=10)
    pyplot.xticks(arange(8), [("%.1f" % ((v + 1) / 2)) + "%" for v in range(8)], fontsize=10)
    pyplot.yticks([0, 2, 4, 6, 8, 10], [1, 4, 16, 64, 256, 1024], fontsize=10)
    pyplot.xlim(-0.8, 7.8)
    pyplot.ylim(-0.5, 10.5)
    pyplot.savefig("./show/supp12.pdf", format="pdf", bbox_inches="tight", dpi=600)
    pyplot.close()


def supp13():
    data = load_data(load_path="./raw/correction_evaluation_4.pkl")

    figure = pyplot.figure(figsize=(10, 5))

    ax1 = pyplot.subplot(1, 2, 1)
    values = array([data[(0.005, 0.990)].tolist(), data[(0.005, 0.999)].tolist(), data[(0.005, 1.000)].tolist()])
    pyplot.title("error rates = 0.5%", fontsize=12)
    a = pyplot.pcolormesh(arange(6), arange(4), values, vmin=0, vmax=60, cmap="rainbow")
    for x in range(5):
        for y in range(3):
            pyplot.text(x + 0.5, y + 0.5, values[y, x], va="center", ha="center", fontsize=9)
    pyplot.xlabel("sequence diversity", fontsize=10)
    pyplot.ylabel("retrieval rate", fontsize=10)
    pyplot.xticks(arange(5) + 0.5, ["1E+" + str(v) for v in arange(1, 6)], fontsize=10)
    pyplot.yticks(arange(3) + 0.5, ["99.0%", "99.9%", "100%"], fontsize=10)
    pyplot.xlim(0, 5)
    pyplot.ylim(0, 3)

    ax2 = pyplot.subplot(1, 2, 2)
    values = array([data[(0.040, 0.990)].tolist(), data[(0.040, 0.999)].tolist(), data[(0.040, 1.000)].tolist()])
    pyplot.title("error rates = 4.0%", fontsize=12)
    pyplot.pcolormesh(arange(6), arange(4), values, vmin=0, vmax=60, cmap="rainbow")
    for x in range(5):
        for y in range(3):
            pyplot.text(x + 0.5, y + 0.5, values[y, x], va="center", ha="center", fontsize=9)
    pyplot.xlabel("sequence diversity", fontsize=10)
    pyplot.ylabel("retrieval rate", fontsize=10)
    pyplot.xticks(arange(5) + 0.5, ["1E+" + str(v) for v in arange(1, 6)], fontsize=10)
    pyplot.yticks(arange(3) + 0.5, ["99.0%", "99.9%", "100%"], fontsize=10)
    pyplot.xlim(0, 5)
    pyplot.ylim(0, 3)

    bar = figure.colorbar(a, ax=[ax1, ax2], shrink=0.75, orientation="horizontal", ticks=[0, 15, 30, 45, 60])
    bar.set_label("minimum reads number to reach the given retrieval rate", fontsize=10)
    bar.ax.tick_params(labelsize=10)

    pyplot.savefig("./show/supp13.pdf", format="pdf", bbox_inches="tight", dpi=600)
    pyplot.close()


def supp14():
    data = load_data(load_path="./raw/correction_evaluation_6.pkl")

    figure = pyplot.figure(figsize=(10, 6))

    ax1 = pyplot.subplot(1, 2, 1)
    pyplot.title("error rates = 0.5%", fontsize=12)
    a = pyplot.pcolormesh(arange(7), arange(6), data[0.005], vmin=0, vmax=20, cmap="rainbow")
    for x in range(6):
        for y in range(5):
            pyplot.text(x + 0.5, y + 0.5, data[0.005][y, x], va="center", ha="center", fontsize=10)
    pyplot.xlabel("sequence diversity", fontsize=10)
    pyplot.ylabel("reads number", fontsize=10)
    pyplot.xticks(arange(6) + 0.5, ["1E+" + str(v) for v in arange(1, 7)], fontsize=10)
    pyplot.yticks(arange(5) + 0.5, [10, 20, 30, 40, 50], fontsize=10)
    pyplot.xlim(0, 6)
    pyplot.ylim(0, 5)

    ax2 = pyplot.subplot(1, 2, 2)
    pyplot.title("error rates = 4.0%", fontsize=12)
    pyplot.pcolormesh(arange(7), arange(6), data[0.040], vmin=0, vmax=20, cmap="rainbow")
    for x in range(6):
        for y in range(5):
            pyplot.text(x + 0.5, y + 0.5, data[0.040][y, x], va="center", ha="center", fontsize=10)
    pyplot.xlabel("sequence diversity", fontsize=10)
    pyplot.ylabel("reads number", fontsize=10)
    pyplot.xticks(arange(6) + 0.5, ["1E+" + str(v) for v in arange(1, 7)], fontsize=10)
    pyplot.yticks(arange(5) + 0.5, [10, 20, 30, 40, 50], fontsize=10)
    pyplot.xlim(0, 6)
    pyplot.ylim(0, 5)

    bar = figure.colorbar(a, ax=[ax1, ax2], shrink=0.75, orientation="horizontal", ticks=[0, 5, 10, 15, 20])
    bar.set_label("maximum frequency of incorrect reads", fontsize=10)
    bar.ax.tick_params(labelsize=10)

    pyplot.savefig("./show/supp14.pdf", format="pdf", bbox_inches="tight", dpi=600)
    pyplot.close()


def supp15():
    data = load_data(load_path="./raw/correction_evaluation_7.pkl")

    book = Workbook()
    sheet = book.create_sheet(title="supp15", index=0)
    sheet.append(["error rate", "SPIDER-WEB (average)", "SPIDER-WEB (median)", "HEDGES (average)", "HEDGES (median)"])
    for value in range(9):
        value_x = data["spiderweb"][value / 200.0]
        value_y = data["hedges"][value / 200.0]
        sheet.append([value / 200.0, mean(value_x), median(value_x), mean(value_y), median(value_y)])

    book.save("./show/supp15.xlsx")


def supp16():
    filter_indices = ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12"]
    coding_graphs = load_data(load_path="./raw/graph_coding.pkl")
    follow_ups = zeros(shape=(12, 3))
    for index, filter_index in enumerate(filter_indices):
        accessor = coding_graphs[filter_index]
        out_degrees, counts = array(list(Counter(where(accessor >= 0)[0]).values())), zeros(shape=(3,), dtype=float)
        for out_degree in [2, 3, 4]:
            counts[out_degree - 2] = len(where(out_degrees == out_degree)[0])
        counts /= sum(counts)
        follow_ups[index] = counts

    figure = pyplot.figure(figsize=(10, 6), tight_layout=True)
    gradient_colors = pyplot.get_cmap("RdYlGn")(linspace(0, 1, 12))
    for index, filter_index in enumerate(filter_indices):
        lengths = linspace(0, 256, 257)
        rate = follow_ups[index][0] * log2(2.0) + follow_ups[index][1] * log2(6.0) + follow_ups[index][2] * log2(24.0)
        pyplot.plot(lengths, lengths * rate, color=gradient_colors[index], lw=3, zorder=2,
                    label="constraint [" + filter_index + "]")
    pyplot.legend(loc="lower right", ncol=3, fontsize=9)
    pyplot.xlabel("length of DNA sequence", fontsize=10)
    pyplot.ylabel("combination size", fontsize=10)
    pyplot.xticks([0, 64, 128, 192, 256], [0, 64, 128, 192, 256], fontsize=10)
    pyplot.yticks([0, 64, 128, 192, 256], ["2^0", "2^64", "2^128", "2^192", "2^256"], fontsize=10)
    pyplot.xlim(0, 256)
    pyplot.ylim(0, 256)
    pyplot.savefig("./show/supp16.pdf", format="pdf", bbox_inches="tight", dpi=600)
    pyplot.close()


def supp17():
    record = load_data(load_path="./raw/capacity_evaluation.pkl")

    data, errors = record["detail"]
    errors = log10(array(errors))
    x = linspace(min(errors), max(errors), 101)
    y = gaussian_kde(errors)(x)
    y = y / sum(y) * 100.0

    v_25, v_50, v_75 = percentile(a=errors, q=[25, 50, 75])
    lower, upper = v_25 - 1.5 * (v_75 - v_25), 1.5 * (v_75 - v_25) + v_75

    figure = pyplot.figure(figsize=(10, 11), tight_layout=True)

    pyplot.subplot(3, 1, 1)

    pyplot.plot(x, y, color="#81B8DF", linewidth=1.5, zorder=4)
    pyplot.fill_between(x, y, zeros_like(y), color="#B1CCDF", alpha=0.75, zorder=3)

    pyplot.vlines(x=v_50, ymin=0, ymax=6.5, color="black", linewidth=1, zorder=2)
    pyplot.vlines(x=v_50, ymin=6.9, ymax=8.5, color="black", linewidth=1, zorder=2)
    pyplot.hlines(y=8.5, xmin=v_50 - 0.3, xmax=v_50 + 0.3, color="black", linewidth=1, zorder=2)
    pyplot.text(x=v_50, y=8.6, s="median", va="bottom", ha="center", fontsize=10, zorder=2)

    pyplot.vlines(x=v_25, ymin=0, ymax=7, color="black", linewidth=1, zorder=2)
    pyplot.vlines(x=v_75, ymin=0, ymax=7, color="black", linewidth=1, zorder=2)
    pyplot.hlines(y=6.7, xmin=v_25, xmax=v_75, color="black", linewidth=1, zorder=2)
    pyplot.text(x=v_75 + 0.1, y=6.7, s="interquartile range", va="center", ha="left", fontsize=10, zorder=2)
    pyplot.vlines(x=lower, ymin=0, ymax=4, color="black", linewidth=1, zorder=2)
    pyplot.hlines(y=3.7, xmin=lower, xmax=min(errors) + 2, color="black", linewidth=1, zorder=2)
    pyplot.text(x=min(errors) + 1.9, y=3.7, s="outlier range", va="center", ha="right", fontsize=10, zorder=2)
    pyplot.vlines(x=upper, ymin=0, ymax=4, color="black", linewidth=1, zorder=2)
    pyplot.hlines(y=3.7, xmin=upper, xmax=max(errors) - 2, color="black", linewidth=1, zorder=2)
    pyplot.text(x=max(errors) - 1.9, y=3.7, s="outlier range", va="center", ha="left", fontsize=10, zorder=2)

    pyplot.xlim(-14.1, -3.9)
    pyplot.ylim(-0.5, 10.5)
    pyplot.ylabel("frequency", fontsize=10)
    pyplot.xticks([-14, -13, -12, -11, -10, -9, -8, -7, -6, -5, -4],
                  ["1E\N{MINUS SIGN}14", "1E\N{MINUS SIGN}13", "1E\N{MINUS SIGN}12", "1E\N{MINUS SIGN}11",
                   "1E\N{MINUS SIGN}10", "1E\N{MINUS SIGN}09", "1E\N{MINUS SIGN}08", "1E\N{MINUS SIGN}07",
                   "1E\N{MINUS SIGN}06", "1E\N{MINUS SIGN}05", "1E\N{MINUS SIGN}04"],
                  fontsize=10)
    pyplot.xlabel("relative error", fontsize=10)
    pyplot.yticks([0, 2, 4, 6, 8, 10], ["0%", "2%", "4%", "6%", "8%", "10%"], fontsize=10)

    pyplot.subplot(3, 1, 2)
    pyplot.scatter(errors, data, color="#81B8DF", edgecolor="black", zorder=3)

    temp_errors, temp_data = [], []
    for value_1, value_2 in sorted(zip(errors, data), key=lambda v: v[0]):
        temp_errors.append(value_1)
        temp_data.append(value_2)
    errors, data = array(temp_errors), array(temp_data)

    a, b = polyfit(errors, data, deg=1)
    estimated = a * errors + b
    bounds = abs(std(errors) * sqrt(1.0 / len(errors)
                                    + (errors - mean(errors)) ** 2 / sum((errors - mean(errors)) ** 2)))
    pyplot.fill_between(errors, estimated - bounds, estimated + bounds, color="#B1CCDF", alpha=0.75, zorder=1)
    x = linspace(min(errors), max(errors), 100)
    y = a * x + b
    pyplot.plot(x, y, color="black", linestyle="--", linewidth=1, zorder=2)

    pyplot.xlabel("relative error", fontsize=10)
    pyplot.ylabel("capacity from SPIDER-WEB", fontsize=10)
    pyplot.xlim(-14.1, -3.9)
    pyplot.ylim(-0.1, 2.1)
    pyplot.xticks([-14, -13, -12, -11, -10, -9, -8, -7, -6, -5, -4],
                  ["1E\N{MINUS SIGN}14", "1E\N{MINUS SIGN}13", "1E\N{MINUS SIGN}12", "1E\N{MINUS SIGN}11",
                   "1E\N{MINUS SIGN}10", "1E\N{MINUS SIGN}09", "1E\N{MINUS SIGN}08", "1E\N{MINUS SIGN}07",
                   "1E\N{MINUS SIGN}06", "1E\N{MINUS SIGN}05", "1E\N{MINUS SIGN}04"],
                  fontsize=10)
    pyplot.yticks([0, 0.5, 1.0, 1.5, 2.0], ["0.0", "0.5", "1.0", "1.5", "2.0"], fontsize=10)
    whole_data = record["extend"]

    pyplot.subplot(3, 1, 3)

    pyplot.fill_between([1.5, 2.5], [-15, -15], [-3, -3], color="#F1F1F1", zorder=0)
    pyplot.fill_between([3.5, 4.5], [-15, -15], [-3, -3], color="#F1F1F1", zorder=0)

    used = []
    for length, data in whole_data.items():
        used.append(log10(data[1]))

    violin = pyplot.violinplot(dataset=used, positions=[0.9, 1.9, 2.9, 3.9, 4.9], showextrema=False, widths=0.35)
    for patch in violin["bodies"]:
        patch.set_edgecolor("#81B8DF")
        patch.set_facecolor("#B1CCDF")
        patch.set_linewidth(1)
        patch.set_alpha(1)

    for index, data in enumerate(used):
        [v_25, v_50, v_75] = percentile(a=data, q=[25, 50, 75])
        pyplot.scatter(x=[index + 0.9], y=[v_50],
                       color="white", edgecolor="#81B8DF", linewidth=1, s=25, zorder=4)
        lower, upper = v_25 - 1.5 * (v_75 - v_25), 1.5 * (v_75 - v_25) + v_75
        pyplot.vlines(x=index + 1.2, ymin=lower, ymax=upper,
                      color="#81B8DF", linewidth=1, zorder=2)
        pyplot.hlines(y=lower, xmin=index + 1.2 - 0.04, xmax=index + 1.2 + 0.04,
                      color="#81B8DF", linewidth=1, zorder=2)
        pyplot.hlines(y=upper, xmin=index + 1.2 - 0.04, xmax=index + 1.2 + 0.04,
                      color="#81B8DF", linewidth=1, zorder=2)
        pyplot.fill_between(x=[index + 1.2 - 0.04, index + 1.2 + 0.04], y1=[v_25, v_25], y2=[v_75, v_75],
                            color="#81B8DF", zorder=3)
        for value in data:
            if value < lower or value > upper:
                pyplot.scatter(x=[index + 1.2], y=[value],
                               color="white", edgecolor="#81B8DF", linewidth=1, s=20, zorder=4)
    pyplot.xlabel("size of adjacency matrix", fontsize=10)
    pyplot.ylabel("relative error", fontsize=10)
    pyplot.xlim(0.5, 5.5)
    pyplot.ylim(-15, -3)
    pyplot.xticks([1, 2, 3, 4, 5],
                  ["4^2", "4^3", "4^4", "4^5", "4^6"],
                  fontsize=10)
    pyplot.yticks([-15, -13, -11, -9, -7, -5, -3],
                  ["1E\N{MINUS SIGN}15", "1E\N{MINUS SIGN}13", "1E\N{MINUS SIGN}11", "1E\N{MINUS SIGN}09",
                   "1E\N{MINUS SIGN}07", "1E\N{MINUS SIGN}05", "1E\N{MINUS SIGN}03"],
                  fontsize=10)

    figure.align_labels()
    figure.text(0.024, 0.99, "a", va="center", ha="center", fontsize=14)
    figure.text(0.024, 0.66, "b", va="center", ha="center", fontsize=14)
    figure.text(0.024, 0.34, "c", va="center", ha="center", fontsize=14)

    pyplot.savefig("./show/supp17.pdf", format="pdf", bbox_inches="tight", dpi=600)
    pyplot.close()


if __name__ == "__main__":
    supp01()
    supp02()
    supp03()
    supp04()
    supp05()
    supp06()
    supp07()
    supp08()
    supp09()
    supp10()
    supp11()
    supp12()
    supp13()
    supp14()
    supp15()
    supp16()
    supp17()
