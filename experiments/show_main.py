from collections import Counter
# noinspection PyPackageRequirements
from matplotlib import pyplot, rcParams, patches
from numpy import load, zeros, array, linspace, arange, argmax, where
from numpy import maximum, min, mean, median, max, sum, log, log2, log10
# noinspection PyPackageRequirements
from openpyxl import Workbook
from scipy.stats import gaussian_kde
from scipy.optimize import curve_fit
from warnings import filterwarnings

from dsw import obtain_vertices
from experiments import local_bio_filters

filterwarnings("ignore")


def main01():
    print("main01.pdf is the \"illustration of SPIDER-WEB\", created by PowerPoint.")


def main02():
    capacities, _ = load(file="./data/capacity_reference.pkl", allow_pickle=True)

    book = Workbook()
    sheet = book.create_sheet(title="main03", index=0)
    sheet.append(["constraint set index", "long homopolymer", "regionalized GC content", "undesired motifs",
                  "capacity"])

    for constraint_index, bio_filter in local_bio_filters.items():
        save_data = [constraint_index]
        row_info = constraint_index.ljust(20)
        if bio_filter.max_homopolymer_runs is not None:
            row_info += " | " + str(bio_filter.max_homopolymer_runs).ljust(20)
            save_data.append(str(bio_filter.max_homopolymer_runs))
        else:
            row_info += " | " + "N/A".ljust(20)
            save_data.append("N/A")

        if bio_filter.gc_range is not None:
            if bio_filter.gc_range[0] != bio_filter.gc_range[1]:
                gc_range = [str(int(bio_filter.gc_range[0] * 100)) + "%", str(int(bio_filter.gc_range[1] * 100)) + "%"]
                save_data.append(gc_range[0] + " ~ " + gc_range[1])
                row_info += " | " + (gc_range[0] + " ~ " + gc_range[1]).ljust(25)
            else:
                row_info += " | " + (str(int(bio_filter.gc_range[0] * 100)) + "%").ljust(25)
                save_data.append(str(int(bio_filter.gc_range[0] * 100)) + "%")
        else:
            row_info += " | " + "N/A".ljust(25)
            save_data.append("N/A")

        if bio_filter.undesired_motifs is not None:
            row_info += " | " + str(bio_filter.undesired_motifs)[1:-1].replace("\"", "").replace("\'", "").ljust(100)
            save_data.append(str(bio_filter.undesired_motifs)[1:-1].replace("\"", "").replace("\'", ""))
        else:
            row_info += " | " + "N/A".ljust(100)
            save_data.append("N/A")

        row_info += " | " + "%.4f" % capacities[int(constraint_index) - 1]
        save_data.append("%.4f" % capacities[int(constraint_index) - 1])

        sheet.append(save_data)

    book.save("./show/main02.xlsx")


def main03():
    display_data = [[[] for _ in range(4)] for _ in range(12)]
    record = load(file="./data/performance_evaluation_1.pkl", allow_pickle=True)
    for algorithm_index, filter_index, _, _, code_rate in record:
        display_data[int(filter_index - 1)][int(algorithm_index - 1)].append(code_rate)
    book = Workbook()
    sheet = book.create_sheet(title="main04", index=0)
    sheet.append(["constraint set index", "DNA Fountain", "Yin-Yang Code", "HEDGES", "SPIDER-WEB"])
    for index, data in enumerate(display_data):
        record = [str(index + 1).zfill(2)]
        for method_index, sub_data in enumerate(data):
            if sum(sub_data) == -len(sub_data):
                record.append("N/A")
            elif max(sub_data) - min(sub_data) < 0.001:
                record.append("%.2f" % mean(sub_data))
            else:
                sub_data = array(sub_data)
                sub_data = sub_data[sub_data > 0]
                record.append("%.2f ~ %.2f" % (min(sub_data), max(sub_data)))
        sheet.append(record)

    book.save("./show/main03.xlsx")


def main04():
    figure = pyplot.figure(figsize=(10, 3), tight_layout=True)
    rcParams["font.family"] = "Times New Roman"
    used_colors = ["#F19D91", "#9EDBE9", "#74CCBE", "#95A2BF"]

    data = load(file="./data/correction_evaluation_1.pkl", allow_pickle=True)

    pyplot.subplot(1, 3, 1)
    for index, (values, _) in enumerate(data.values()):
        success_rate = sum(values[:, 4]) / 10000.0
        pyplot.bar([index], [success_rate], width=0.5, color=used_colors[index], edgecolor="black", linewidth=0.75)
        pyplot.text(index, success_rate + 0.01, "%.1f" % (success_rate * 100) + "%",
                    va="bottom", ha="center", fontsize=9)
    pyplot.xlabel("error rate", fontsize=10)
    pyplot.ylabel("correction rate", fontsize=10)
    pyplot.xlim(-0.5, 3.5)
    pyplot.ylim(0.0, 1.0)
    pyplot.xticks([0, 1, 2, 3], ["0.5%", "1.0%", "2.0%", "3.0%"], fontsize=10)
    pyplot.yticks([0.0, 0.2, 0.4, 0.6, 0.8, 1.0], ["0%", "20%", "40%", "60%", "80%", "100%"], fontsize=10)

    pyplot.subplot(1, 3, 2)
    for index, (values, _) in enumerate(data.values()):
        counts = values[:, 3]
        counts = counts[counts > 0]
        violin = pyplot.violinplot([log2(counts)], positions=[index], widths=0.5, showextrema=False)
        for patch in violin["bodies"]:
            patch.set_edgecolor("black")
            patch.set_facecolor("#EEEEEE")
            patch.set_linewidth(0.75)
            patch.set_alpha(1)

        label = "median " + ("  " if index < 3 else "") + str(int(median(counts)))
        label += " candidates" if int(median(counts)) > 1 else " candidate"
        pyplot.scatter([index], [median(log2(counts))], color=used_colors[index], edgecolor="black",
                       linewidth=0.75, label=label)
    pyplot.legend(loc="upper left", fontsize=9)
    pyplot.xlabel("error rate", fontsize=10)
    pyplot.ylabel("candidate number from correction", fontsize=10)
    pyplot.xlim(-0.5, 3.5)
    pyplot.ylim(0, 10)
    pyplot.xticks([0, 1, 2, 3], ["0.5%", "1.0%", "2.0%", "3.0%"], fontsize=10)
    pyplot.yticks([0, 2, 4, 6, 8, 10], [1, 4, 16, 64, 256, 1024], fontsize=10)

    pyplot.subplot(1, 3, 3)
    speeds = []
    for index, (_, time) in enumerate(data.values()):
        speed = 200.0 / time
        speeds.append(speed)
        pyplot.text(index, log10(speed) + 0.2, str(int(speed)), va="bottom", ha="center", fontsize=9, zorder=1)
        pyplot.scatter([index], [log10(speed)], color=used_colors[index], edgecolor="black", linewidth=0.75, zorder=2)
    pyplot.plot(arange(4), log10(array(speeds)), color="black", linewidth=0.75, zorder=1)
    pyplot.xlabel("error rate", fontsize=10)
    pyplot.ylabel("average correction speed (bases / second)", fontsize=10)
    pyplot.xlim(-0.5, 3.5)
    pyplot.ylim(2, 7)
    pyplot.xticks([0, 1, 2, 3], ["0.5%", "1.0%", "2.0%", "3.0%"], fontsize=10)
    pyplot.yticks([2, 3, 4, 5, 6, 7], ["1E+2", "1E+3", "1E+4", "1E+5", "1E+6", "1E+7"], fontsize=10)

    figure.align_labels()
    figure.text(0.022, 0.99, "a", va="center", ha="center", fontsize=14)
    figure.text(0.352, 0.99, "b", va="center", ha="center", fontsize=14)
    figure.text(0.679, 0.99, "c", va="center", ha="center", fontsize=14)
    pyplot.savefig("./show/main04.pdf", format="pdf", bbox_inches="tight", dpi=600)
    pyplot.close()


def main05():
    figure = pyplot.figure(figsize=(10, 3), tight_layout=True)
    rcParams["font.family"] = "Times New Roman"
    used_colors = ["#F19D91", "#9EDBE9", "#74CCBE", "#95A2BF"]

    data = load(file="./data/correction_evaluation_2.pkl", allow_pickle=True)

    pyplot.subplot(1, 3, 1)
    locations, error_rates = [2, 1, 0, 0], [0.03, 0.02, 0.01, 0.005]
    for index, error_rate in enumerate(error_rates):
        values = data[error_rate]
        counts = linspace(1, max(values), max(values))
        proportions = gaussian_kde(dataset=values, bw_method=0.5)(counts)
        proportions = proportions / proportions[locations[index]] * (Counter(values)[locations[index] + 1] / 10000.0)
        label = "%.1f" % (error_rates[3 - index] * 100) + "% error rate"
        pyplot.plot(log2(counts), proportions, color=used_colors[3 - index], linewidth=3)
        pyplot.hlines(-1, 0, 0, color=used_colors[index], linewidth=3, label=label, zorder=3)
    pyplot.legend(loc="upper right", fontsize=9)
    pyplot.xlabel("pretreatment-free reads", fontsize=10)
    pyplot.ylabel("proportion", fontsize=10)
    pyplot.xlim(0, 5)
    pyplot.ylim(0, 0.8)
    pyplot.xticks([0, 1, 2, 3, 4, 5], [1, 2, 4, 8, 16, 32], fontsize=10)
    pyplot.yticks([0, 0.2, 0.4, 0.6, 0.8, 1.0], ["0%", "20%", "40%", "60%", "80%", "100%"], fontsize=10)

    data = load(file="./data/correction_evaluation_3.pkl", allow_pickle=True)

    pyplot.subplot(1, 3, 2)
    location, right_color, wrong_color = 1, "#EA8250", "#BBBBBB"
    for count in [5, 4, 3, 2, 1]:
        number = len(where(data[:, 1] == count)[0])
        if location + number < 10000:
            ranges = [location, location + number]
            pyplot.fill_between(log10(ranges), 0, count, color=right_color, linewidth=0, zorder=1)
        elif location < 10000 and location + number > 10000:
            ranges = [location, 10000]
            pyplot.fill_between(log10(ranges), 0, count, color=right_color, linewidth=0, zorder=1)
            ranges = [10001, location + number]
            pyplot.fill_between(log10(ranges), 0, count, color=wrong_color, linewidth=0, zorder=1)
        else:
            ranges = [location, location + number]
            pyplot.fill_between(log10(ranges), 0, count, color=wrong_color, linewidth=0, zorder=1)
        location += number
    wrong_counts = [0, 0]
    for location, (flag, count) in enumerate(data):
        if not flag:
            if count == 2:
                wrong_counts[0] += 1
            else:
                wrong_counts[1] += 1
        if location == 10000:
            break
    stop_2, rate = len(where(data[:, 1] >= 2)[0]), sum(data[:10000, 0]) / 100.0
    pyplot.fill_between(log10([stop_2 - wrong_counts[0], stop_2]), 0, 2, color=wrong_color, linewidth=0, zorder=2)
    pyplot.fill_between(log10([10000 - wrong_counts[1], 10000]), 0, 1, color=wrong_color, linewidth=0, zorder=2)
    pyplot.fill_between(log10([10000, 20000 - sum(data[:10000, 0])]), 0, 1, color=right_color, linewidth=0, zorder=2)
    pyplot.annotate(s="", xy=(0, 6), xytext=(4, 6),
                    arrowprops=dict(arrowstyle="<|-|>", color="black", shrinkA=0, shrinkB=0, lw=0.75), zorder=2)
    pyplot.text(2, 6, "overall success rate\n%.2f" % rate + "%",
                bbox=dict(facecolor="white", edgecolor="white"), va="center", ha="center", fontsize=9)
    handles = [patches.Patch(facecolor=right_color, label="right candidate"),
               patches.Patch(facecolor=wrong_color, label="wrong candidate")]
    pyplot.legend(handles=handles, loc="upper left", fontsize=9)
    pyplot.vlines(log10(10000), 0, 10, color="black", linestyle="--", linewidth=0.75, zorder=3)
    pyplot.xlabel("candidate priority", fontsize=10)
    pyplot.ylabel("occurrence frequency", fontsize=10)
    pyplot.xlim(0, 6)
    pyplot.ylim(0, 10)
    pyplot.xticks([0, 2, 4, 6], ["1E+0", "1E+2", "1E+4", "1E+6"], fontsize=10)
    pyplot.yticks([0, 2, 4, 6, 8, 10], [0, 2, 4, 6, 8, 10], fontsize=10)

    data = load(file="./data/correction_evaluation_4.pkl", allow_pickle=True)

    pyplot.subplot(1, 3, 3)
    processes = zeros(shape=(21, 3))
    for index, values in enumerate(data[0.02]):
        bounds = array([min(values[:, 0]), mean(values[:, 0]), max(values[:, 0])])
        shown_values = (5.0 - maximum(log10(100000.0 * (1 - bounds)), 0)) / 5.0
        processes[index + 1] = shown_values
    pyplot.fill_between(arange(21), processes[:, 0], processes[:, 2], color=used_colors[2], label="range")
    pyplot.plot(arange(21), processes[:, 1], color="black", linewidth=1, linestyle="--", label="average level")
    location = (5.0 - log10(100000.0 * (1 - rate * 0.01))) / 5.0
    pyplot.annotate(s="", xy=(7.0, location - 0.1), xytext=(5.0, location),
                    arrowprops=dict(arrowstyle="-|>", color="black", shrinkA=1, shrinkB=2, lw=1), zorder=2)
    pyplot.text(7.0, location - 0.1, "b (" + ("%.2f" % rate) + "%)", fontsize=9, va="top", ha="left")
    pyplot.legend(loc="upper left", fontsize=9)
    pyplot.xlabel("pretreatment-free depth", fontsize=10)
    pyplot.ylabel("overall success rate", fontsize=10)
    pyplot.xlim(0, 20)
    pyplot.ylim(0, 1.0)
    pyplot.xticks([0, 5, 10, 15, 20], [0, 5, 10, 15, 20], fontsize=10)
    pyplot.yticks([0, 0.2, 0.4, 0.6, 0.8, 1.0],
                  ["00.00%", "90.00%", "99.00%", "99.90%", "99.99%", "perfect "], fontsize=10)

    figure.align_labels()
    figure.text(0.021, 0.99, "a", va="center", ha="center", fontsize=14)
    figure.text(0.374, 0.99, "b", va="center", ha="center", fontsize=14)
    figure.text(0.680, 0.99, "c", va="center", ha="center", fontsize=14)

    pyplot.savefig("./show/main05.pdf", format="pdf", bbox_inches="tight", dpi=600)
    pyplot.close()


def main06():
    # bit length or key strength is 256.
    # British Standards Institution (2020). Cryptographic Mechanisms: Recommendations and Key Lengths.
    bit_length = 256

    def estimated_equation(x, a):
        return a ** x

    reconstructions = load(file="./data/reconstruction_evaluation.npy")
    capacities, _ = load("./data/capacity_reference.pkl", allow_pickle=True)
    accessors = load("./data/coding_graphs.pkl", allow_pickle=True)

    filter_indices = ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12"]
    gradient_colors = pyplot.get_cmap(name="rainbow")(linspace(0, 1, 12))

    figure = pyplot.figure(figsize=(10, 3), tight_layout=True)
    rcParams["font.family"] = "Times New Roman"

    pyplot.subplot(1, 2, 1)

    collected_data = zeros(shape=(12, 100))
    for index, data in enumerate(reconstructions):
        number = len(obtain_vertices(accessors[str(index + 1).zfill(2)]))
        for iteration_index, iteration_data in enumerate(data):
            if number > max(iteration_data):
                sequences = linspace(0, 10000, 10001)
                parameter = curve_fit(estimated_equation, xdata=iteration_data, ydata=sequences)[0][0]
                log_sequence_number = number * log(parameter)
            else:
                log_sequence_number = log(argmax(iteration_data))
            value = (log_sequence_number + log(100) + log(capacities[index]) - log(8)) / log(1024.0)
            collected_data[index, iteration_index] = value

    for index in range(12):
        if max(collected_data[index]) - min(collected_data[index]) > 0.05:
            pyplot.plot([index - 0.25, index + 0.25], [min(collected_data[index]), max(collected_data[index])],
                        color=gradient_colors[index], linewidth=2)
            pyplot.plot([index + 0.25, index - 0.25], [min(collected_data[index]), max(collected_data[index])],
                        color=gradient_colors[index], linewidth=2)
        else:
            pyplot.hlines(median(collected_data[index]), index - 0.25, index + 0.25,
                          color=gradient_colors[index], linewidth=2)

    pyplot.xlabel("constraint set index", fontsize=10)
    pyplot.xticks(range(12), filter_indices, fontsize=10)
    pyplot.xlim(-0.5, 11.5)
    pyplot.ylabel("transmitted file size", fontsize=10)
    pyplot.yticks([0, 1, 2, 3, 4], ["B", "KB", "MB", "GB", "TB"], fontsize=10)
    pyplot.ylim(0, 4)

    pyplot.subplot(1, 2, 2)
    follow_ups = zeros(shape=(12, 3))  # [2, 3, 4]
    for index, filter_index in enumerate(filter_indices):
        accessor = accessors[filter_index]
        out_degrees, counts = array(list(Counter(where(accessor >= 0)[0]).values())), zeros(shape=(3,), dtype=float)
        for out_degree in [2, 3, 4]:
            counts[out_degree - 2] = len(where(out_degrees == out_degree)[0])
        counts /= sum(counts)
        follow_ups[index] = counts

    pyplot.hlines(bit_length, 0, 320, color="silver", linewidth=0.75, linestyle="--", zorder=1)
    pyplot.text(4, 256 + 4, "AES-256", va="bottom", ha="left", fontsize=10)
    for index, filter_index in enumerate(filter_indices):
        lengths = linspace(0, 320, 321)
        rate = follow_ups[index][0] * log2(2.0) + follow_ups[index][1] * log2(6.0) + follow_ups[index][2] * log2(24.0)
        pyplot.plot(lengths, lengths * rate, color=gradient_colors[index], alpha=0.5, linewidth=2, zorder=2)

        reference_length = bit_length / rate
        pyplot.vlines(reference_length, 0, bit_length,
                      color="silver", linewidth=0.75, linestyle="--", zorder=1)
        if index > 0:
            pyplot.scatter(reference_length, bit_length,
                           color=gradient_colors[index], edgecolor="black", s=20, zorder=4,
                           label="[" + str(filter_index) + "]   %.1f" % reference_length)
        else:
            pyplot.scatter(reference_length, bit_length,
                           color=gradient_colors[index], edgecolor="black", s=20, zorder=4,
                           label="[" + str(filter_index) + "] %.1f" % reference_length)

    pyplot.legend(loc="upper right", ncol=3, fontsize=9)
    pyplot.xlabel("DNA string length", fontsize=10)
    pyplot.xlim(0, 320)
    pyplot.xticks([0, 64, 128, 192, 256, 320], ["0nt", "64nt", "128nt", "192nt", "256nt", "320nt"], fontsize=10)
    pyplot.ylabel("equivalent key strength", fontsize=10)
    pyplot.ylim(0, 512)
    pyplot.yticks([0, 128, 256, 384, 512], [0, 128, 256, 384, 512], fontsize=10)

    figure.align_labels()
    figure.text(0.020, 0.98, "a", va="center", ha="center", fontsize=14)
    figure.text(0.506, 0.98, "b", va="center", ha="center", fontsize=14)

    pyplot.savefig("./show/main06.pdf", format="pdf", bbox_inches="tight", dpi=600)
    pyplot.close()


if __name__ == "__main__":
    main01()
    main02()
    main03()
    main04()
    main05()
    main06()
