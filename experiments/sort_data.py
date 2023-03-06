from hashlib import md5
from numpy import random, array, ones, linspace, repeat, expand_dims, vstack, log10, ceil, where
from openpyxl import Workbook
from os.path import exists, getsize

from dsw import obtain_vertices

from experiments import local_bio_filters, load_data


def supplementary_data():
    if not exists(path="./data/supplementary data.xlsx"):
        book = Workbook()

        sheet = book.create_sheet(title="Figure 1")
        sheet.append(["Illustration does not exist supplementary data"])

        sheet = book.create_sheet(title="Figure 2a")
        sheet.append(["Illustration does not exist supplementary data"])

        sheet = book.create_sheet(title="Figure 2b,4a,S4,S5")
        sheet.append(["error rate", "error number", "detection number", "detection flag",
                      "candidate number (search only)", "candidate number (combined)", "correction flag"])
        record = load_data(load_path="./raw/correction_evaluation_1.pkl")
        for error_rate, data in record.items():
            for values in data[0]:
                sheet.append([error_rate, int(error_rate * 200), values[0], values[0] == int(error_rate * 200),
                              values[2], values[3], values[3] == 1])

        sheet = book.create_sheet(title="Figure 2c")
        sheet.append(["error rate", "error number", "test index", "reads number"])
        record = load_data(load_path="./raw/correction_evaluation_2.pkl")
        for error_rate, values in record.items():
            if error_rate <= 0.04:
                for index, value in enumerate(values):
                    sheet.append([error_rate, int(error_rate * 200), index + 1, value])

        sheet = book.create_sheet(title="Figure 3a,3b")
        sheet.append(["error rate", "reads number", "maximum loss", "smallest gap between correct-incorrect reads"])
        record = load_data(load_path="./raw/correction_evaluation_3.pkl")
        loss_data, gap_data = record["loss"], record["gap"]
        for idx_1, error_rate in enumerate(linspace(0.005, 0.040, 8)):
            for idx_2, reads_number in enumerate(linspace(5, 50, 10, dtype=int)):
                sheet.append([error_rate, reads_number, loss_data[idx_1, idx_2], gap_data[idx_1, idx_2]])

        sheet = book.create_sheet(title="Figure 3c")
        sheet.append(["error rate", "retrieval rate", "sequence diversity", "minimum reads number"])
        record = load_data("./raw/correction_evaluation_5.pkl")
        for error_rate in [0.040, 0.005]:
            for retrieval_rate in [1.000, 0.999, 0.990]:
                for x_value, y_value in zip(*record[(error_rate, retrieval_rate)]):
                    sheet.append([error_rate, retrieval_rate, x_value, y_value])

        sheet = book.create_sheet(title="Figure 3d")
        sheet.append(["reads number",
                      "non-blocking reads threshold (4.0% | 1E+12)",
                      "non-blocking reads threshold (4.0% | 1E+09)",
                      "non-blocking reads threshold (4.0% | 1E+06)",
                      "non-blocking reads threshold (0.5% | 1E+12)",
                      "non-blocking reads threshold (0.5% | 1E+09)",
                      "non-blocking reads threshold (0.5% | 1E+06)"])

        # create by symbolic regression technique (4.0%).
        def function_4_0(diversity, depth):
            return log10(depth + 1) * log10(diversity) / 0.346

        # create by symbolic regression technique (0.5%).
        def function_0_5(diversity, depth):
            return log10(depth + 1) * log10(diversity) / 1.601

        for r in linspace(1, 280, 280):
            v_1, v_2, v_3 = function_4_0(1e12, r), function_4_0(1e9, r), function_4_0(1e6, r)
            v_4, v_5, v_6 = function_0_5(1e12, r), function_0_5(1e9, r), function_0_5(1e6, r)
            sheet.append([r,
                          v_1 if v_1 < r else "N/A", v_2 if v_2 < r else "N/A", v_3 if v_3 < r else "N/A",
                          v_4 if v_4 < r else "N/A", v_5 if v_5 < r else "N/A", v_6 if v_6 < r else "N/A"])

        sheet = book.create_sheet(title="Figure 4a")
        sheet.append(["error rate", "total time cost (10,000 samples; second)", "average speed (base / second)",
                      "performance (compared with high-throughput sequencing; 1 base / second)",
                      "performance (compared with single-molecule sequencing; 450 base / second)"])
        record = load_data(load_path="./raw/correction_evaluation_1.pkl")
        for error_rate, data in record.items():
            if error_rate <= 0.04:
                speed = 200.0 * len(data[0]) / data[1]
                sheet.append([error_rate, len(data[0]), speed, speed / 1.0, speed / 450.0])

        sheet = book.create_sheet(title="Figure 4b")
        sheet.append(["sequence diversity",
                      "approximated operations of conventional pipeline",
                      "approximated operations of our pipeline (4.0% of edit errors)",
                      "approximated operations of our pipeline (0.5% of edit errors"])

        def previous(diversity):
            return 200 * (diversity ** 2) + 48000200 * diversity

        def propose40(diversity):
            a = 7267.904 * diversity * (log10(diversity) ** 2)
            b = 76672.229 * diversity
            return a + b

        def propose05(diversity):
            a = 183.178 * diversity * (log10(diversity) ** 2)
            b = 50508.157 * diversity
            return a + b

        for sequence_diversity in 10 ** linspace(0, 20, 21):
            sheet.append([sequence_diversity, previous(sequence_diversity),
                          propose40(sequence_diversity), propose05(sequence_diversity)])

        sheet = book.create_sheet(title="Table S1")
        sheet.append(["set index", "homopolymer run-length", "regionalized GC content", "undesired motifs", "capacity"])
        record, _ = load_data(load_path="./raw/capacity_reference.pkl")
        for constraint_index, bio_filter in local_bio_filters.items():
            save_data = [constraint_index]
            if bio_filter.max_homopolymer_runs is not None:
                save_data.append(str(bio_filter.max_homopolymer_runs))
            else:
                save_data.append("N/A")

            if bio_filter.gc_range is not None:
                if bio_filter.gc_range[0] != bio_filter.gc_range[1]:
                    gc_range = [str(int(bio_filter.gc_range[0] * 100)) + "%",
                                str(int(bio_filter.gc_range[1] * 100)) + "%"]
                    save_data.append(gc_range[0] + " ~ " + gc_range[1])
                else:
                    save_data.append(str(int(bio_filter.gc_range[0] * 100)) + "%")
            else:
                save_data.append("N/A")

            if bio_filter.undesired_motifs is not None:
                save_data.append(str(bio_filter.undesired_motifs)[1:-1].replace("\"", "").replace("\'", ""))
            else:
                save_data.append("N/A")

            save_data.append("%.4f" % record[int(constraint_index) - 1])

            sheet.append(save_data)

        sheet = book.create_sheet(title="Figure S1")
        sheet.append(["set index", "generation runtime (seconds)"])
        record = load_data(load_path="./raw/generation_evaluation.pkl")["time"]
        for index in range(12):
            sheet.append([str(index + 1).zfill(2), record[index]])

        sheet = book.create_sheet(title="Figure S2")
        record = load_data(load_path="./raw/generation_evaluation.pkl")["change"]
        sheet.append(["set index", "remain vertices after screening"])
        for index in range(12):
            sheet.append([str(index + 1).zfill(2), record[str(index + 1).zfill(2)][0]])

        sheet = book.create_sheet(title="Table S2")
        sheet.append(["trimming cycle",
                      "set 01", "set 02", "set 03", "set 04", "set 05", "set 06",
                      "set 07", "set 08", "set 09", "set 10", "set 11", "set 12"])
        record = load_data(load_path="./raw/generation_evaluation.pkl")["change"]
        for cycle_index in range(13):
            values = [cycle_index]
            for set_index in range(12):
                numbers = record[str(set_index + 1).zfill(2)][1]
                if len(numbers) > cycle_index:
                    values.append(numbers[cycle_index])
                else:
                    values.append("N/A")
            sheet.append(values)

        sheet = book.create_sheet(title="Table S3")
        sheet.append(["param index", "c in DNA Fountain", "delta in DNA Fountain", "rule index in Yin-Yang Code",
                      "pattern in HEDGES", "correct penalty in HEDGES"])
        random_seed, param_number = 2021, 100
        random.seed(random_seed)
        fountain_params = random.uniform(low=array([0.1, 0.01]), high=array([1, 0.1]), size=(param_number, 2)).tolist()
        random.seed(random_seed)
        total_indices = array(list(range(6144)))
        random.shuffle(total_indices)
        yinyang_params = total_indices[:param_number]
        yinyang_params = yinyang_params.tolist()
        hedges_params = []
        correct_penalties = [-0.035, -0.082, -0.127, -0.229, -0.265, -0.324]
        for pattern_index, correct_penalty in enumerate(correct_penalties):
            hedges_params.append([pattern_index + 1, correct_penalty])
        hedges_params = array(hedges_params).tolist()
        for param_index in range(param_number):
            info = [param_index + 1]
            for value in fountain_params[param_index]:
                info.append(value)
            info.append(yinyang_params[param_index])
            if param_index < 6:
                info.append(int(hedges_params[param_index][0]))
                info.append(hedges_params[param_index][1])
            else:
                info.append("N/A")
                info.append("N/A")
            sheet.append(info)

        sheet = book.create_sheet(title="Table S4")
        sheet.append(["param index", *["set " + str(idx + 1).zfill(2) for idx in range(12)]])
        random_seed, param_number, spiderweb_params = 2021, 100, []
        record = load_data(load_path="./raw/graph_coding.pkl")
        for filter_index in ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12"]:
            vertices = obtain_vertices(accessor=record[filter_index])
            random.seed(random_seed)
            random.shuffle(vertices)
            chosen_indices = vertices[:param_number]
            spiderweb_params.append(chosen_indices)
        spiderweb_params = array(spiderweb_params).T.tolist()
        for param_index in range(param_number):
            info = [param_index + 1]
            for vertex_index in spiderweb_params[param_index]:
                info.append(vertex_index)
            sheet.append(info)

        sheet = book.create_sheet(title="Table S5,S6,S7")
        sheet.append(["set index", "parameter index", "data index",
                      "code rate of DNA Fountain", "code rate of Yin-Yang Code", "code rate of HEDGES",
                      "code rate of SPIDER-WEB"])
        record = load_data(load_path="./raw/coding_evaluation_1.pkl")
        column1 = repeat(linspace(1, 12, 12), axis=0, repeats=10000)
        column2 = repeat(expand_dims(repeat(linspace(1, 100, 100), axis=0, repeats=100), axis=0), axis=0,
                         repeats=12).reshape(-1)
        column3 = repeat(expand_dims(linspace(1, 100, 100), axis=0), axis=0, repeats=1200).reshape(-1)
        columns = vstack((column1, column2, column3,
                          -ones(shape=(120000,)), -ones(shape=(120000,)),
                          -ones(shape=(120000,)), -ones(shape=(120000,)))).T
        for index in range(4):
            data = record[where(record[:, 0] == index + 1)]
            if len(data) == 120000:
                columns[:, index + 3] = data[:, -1]
            else:
                for location in range(12):
                    values = data[location * 600: (location + 1) * 600, -1]
                    columns[location * 10000: location * 10000 + 600, index + 3] = values
        for column in columns:
            code_rates = []
            for index in range(4):
                if column[index + 3] > 0:
                    code_rates.append(column[index + 3])
                else:
                    code_rates.append("N/A")
            sheet.append([str(int(column[0])).zfill(2), int(column[1]), int(column[2]), *code_rates])

        sheet = book.create_sheet(title="Figure S3")
        sheet.append(["set index", "capacity", "test index", "code rate"])
        filter_indices = ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12"]
        capacities, _ = load_data(load_path="./raw/capacity_reference.pkl")
        code_results = load_data(load_path="./raw/coding_evaluation_2.pkl")[0]
        for index, filter_index in enumerate(filter_indices):
            capacity, code_rates = capacities[index], code_results[index]
            for test_index, code_rate in enumerate(code_rates):
                sheet.append([filter_index, capacity, test_index + 1, code_rate])

        sheet = book.create_sheet(title="Figure S6")
        sheet.append(["error rate", "retrieval rate", "sequence diversity", "minimum reads number"])
        record = load_data(load_path="./raw/correction_evaluation_4.pkl")
        for error_rate in [0.040, 0.005]:
            for retrieval_rate in [1.000, 0.999, 0.990]:
                for index, sequence_diversity in enumerate([1e1, 1e2, 1e3, 1e4, 1e5]):
                    numbers = record[(error_rate, retrieval_rate)]
                    sheet.append([error_rate, retrieval_rate, sequence_diversity, numbers[index]])

        sheet = book.create_sheet(title="Figure S7")
        sheet.append(["error rate", "reads number", "sequence diversity", "non-blocking reads threshold"])
        record = load_data(load_path="./raw/correction_evaluation_6.pkl")
        for error_rate in [0.040, 0.005]:
            for index_1, reads_number in enumerate([10, 20, 30, 40, 50]):
                for index_2, sequence_diversity in enumerate([1e1, 1e2, 1e3, 1e4, 1e5]):
                    sheet.append([error_rate, reads_number, sequence_diversity, record[error_rate][index_1, index_2]])

        sheet = book.create_sheet(title="Table S8")
        sheet.append(["error rate", "test index",
                      "vertex access frequency of SPIDER-WEB", "vertex access frequency of HEDGES"])
        record = load_data(load_path="./raw/correction_evaluation_7.pkl")
        for value in range(9):
            value_x = record["spiderweb"][value / 200.0]
            value_y = record["hedges"][value / 200.0]
            for test_index, (x, y) in enumerate(zip(value_x, value_y)):
                sheet.append([value / 200.0, test_index + 1, x, y])

        sheet = book.create_sheet(title="Figure S8")
        sheet.append(["matrix size", "test index", "capacity", "relative error"])
        record = load_data(load_path="./raw/capacity_evaluation.pkl")
        for size, data in record["extend"].items():
            for test_index, (capacity, error) in enumerate(zip(data[0], data[1])):
                sheet.append([str(int(4 ** size)) + " x " + str(int(4 ** size)), test_index + 1, capacity, error])

        book.save("./data/supplementary data.xlsx")


if __name__ == "__main__":
    supplementary_data()

    # do it after coloring.
    md5_hash = md5()
    with open("./data/supplementary data.xlsx", "rb") as f:
        for byte_block in iter(lambda: f.read(4096), b""):
            md5_hash.update(byte_block)
    file_size = getsize("./data/supplementary data.xlsx")
    print("./data/supplementary data.xlsx", (md5_hash.hexdigest()).upper(), ceil(file_size / 1024).astype(int))
